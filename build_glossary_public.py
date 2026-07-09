# -*- coding: utf-8 -*-
"""G検定 用語集 xlsx 生成スクリプト（GitHub Pages公開版）。

public_data.json（659語、うち84語は市販参考書の言い回しと一致していた説明文を
独自表現に書き直したもの）を読み込んでxlsxを生成する。用語名・英語名自体は
一般的な技術用語であり著作物ではないため変更していない。書き直し対象は
desc（説明文）が参考書の文章表現に近すぎた語のみ。
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("public_data.json", encoding="utf-8") as f:
    DATA = [tuple(row) for row in json.load(f)]

# 重要度の補正（JDLAシラバス2024の傾向: 生成AI関連を昇格、古典AI固有事例を降格）
IMPORTANCE_OVERRIDE = {
    "プロンプト": "◎", "ChatGPT": "◎", "RLHF": "◎", "ファウンデーションモデル": "◎",
    "GPT-3／GPT-4": "◎", "バイアス・バリアンス": "◎", "正解率": "◎", "TF-IDF": "◎",
    "Cycプロジェクト": "△", "東ロボくん": "△", "イライザ": "△", "セマンティックWeb": "△",
    "意味ネットワーク": "△", "ローブナーコンテスト": "△",
}

# 列定義
HEADERS = ["No", "カテゴリ", "用語", "英語/略語", "意味・説明", "補足・関連用語", "重要度", "チェック"]
WIDTHS  = [5,   22,        26,    26,         52,           30,             8,        8]

wb = Workbook()
ws = wb.active
ws.title = "用語集"

FONT = "Meiryo UI"
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
cell_font   = Font(name=FONT, size=10)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill = PatternFill("solid", fgColor="EEF3F8")

# 重要度の色
imp_fill = {"◎": PatternFill("solid", fgColor="FFD9D9"),
            "○": PatternFill("solid", fgColor="FFF2CC"),
            "△": PatternFill("solid", fgColor="E2EFDA")}

# ヘッダ
ws.append(HEADERS)
for c, w in zip(ws[1], WIDTHS):
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

# データ
for i, (cat, term, en, desc, note, imp) in enumerate(DATA, start=1):
    imp = IMPORTANCE_OVERRIDE.get(term, imp)
    row = [i, cat, term, en, desc, note, imp, ""]
    ws.append(row)
    r = ws.max_row
    for j, c in enumerate(ws[r], start=1):
        c.font = cell_font
        c.border = border
        c.alignment = Alignment(vertical="center",
                                horizontal="center" if j in (1, 7, 8) else "left",
                                wrap_text=(j == 5))
        if i % 2 == 0:
            c.fill = alt_fill
    ws.cell(row=r, column=7).fill = imp_fill.get(imp, PatternFill())

# 列幅
for idx, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = w

# ヘッダ固定・オートフィルタ
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{ws.max_row}"
ws.row_dimensions[1].height = 22

out = "gkentei_glossary_public.xlsx"
wb.save(out)
print("saved:", out, "/", len(DATA), "terms")
