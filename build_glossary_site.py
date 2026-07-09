# -*- coding: utf-8 -*-
"""G検定 用語集 xlsx -> 単一HTML検索サイト（GitHub Pages公開版）を生成する。

gkentei_glossary_public.xlsx（市販参考書と表現が一致していた84語のdescを
書き直したもの）を読み込む。黒本PDF・pagemap.json・pagetext.json は同梱しない。
"""
import json
import os
import openpyxl

XLSX = "gkentei_glossary_public.xlsx"
OUT = "glossary.html"

ws = openpyxl.load_workbook(XLSX).active
records = []
for r in range(2, ws.max_row + 1):
    no, cat, term, en, desc, note, imp, _ = [ws.cell(r, c).value for c in range(1, 9)]
    if not term:
        continue
    records.append({"id": no, "cat": cat or "", "term": term or "", "en": en or "",
                    "desc": desc or "", "note": note or "", "imp": imp or ""})

def catkey(x):
    head = x.split(".")[0]
    return float(head) if head.replace(".", "").isdigit() else 999.0

cats = sorted({r["cat"] for r in records}, key=catkey)
data_json = json.dumps(records, ensure_ascii=False)
cats_json = json.dumps(cats, ensure_ascii=False)

HTML = r"""<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>G検定 用語集 検索</title>
<style>
:root{--bg:#0f172a;--panel:#1e293b;--panel2:#273449;--line:#334155;--text:#e2e8f0;
  --muted:#94a3b8;--accent:#38bdf8;--accent2:#0ea5e9;--imp-a:#f87171;--imp-b:#fbbf24;--imp-c:#34d399;}
*{box-sizing:border-box}
body{margin:0;font-family:"Meiryo UI","Segoe UI",system-ui,sans-serif;background:var(--bg);color:var(--text);line-height:1.6}
header{position:sticky;top:0;z-index:10;background:rgba(15,23,42,.96);backdrop-filter:blur(6px);
  border-bottom:1px solid var(--line);padding:14px 20px}
h1{margin:0 0 10px;font-size:18px;display:flex;align-items:center;gap:10px}
h1 a{color:inherit;text-decoration:none}
h1 a:hover{color:var(--accent)}
h1 .badge{font-size:11px;background:var(--accent2);color:#02283a;padding:2px 8px;border-radius:10px;font-weight:700}
.controls{display:flex;flex-wrap:wrap;gap:10px;align-items:center}
#q{flex:1 1 320px;min-width:220px;padding:10px 14px;font-size:15px;border-radius:10px;
  border:1px solid var(--line);background:var(--panel);color:var(--text);outline:none}
#q:focus{border-color:var(--accent)}
select{padding:9px 10px;border-radius:9px;border:1px solid var(--line);background:var(--panel);color:var(--text);font-size:13px}
.imps{display:flex;gap:6px}
.imps label{display:flex;align-items:center;gap:4px;background:var(--panel);border:1px solid var(--line);
  padding:7px 10px;border-radius:9px;font-size:13px;cursor:pointer;user-select:none}
.imps input{accent-color:var(--accent)}
.count{color:var(--muted);font-size:13px;padding:8px 20px}
.layout{display:flex;height:calc(100vh - 132px)}
.list{flex:1 1 0;overflow-y:auto;padding:8px 12px 60px}
.detail{flex:0 0 380px;border-left:1px solid var(--line);overflow-y:auto;padding:20px;background:var(--panel)}
.card{background:var(--panel);border:1px solid var(--line);border-radius:10px;padding:12px 14px;margin:8px 0;cursor:pointer;transition:.12s}
.card:hover{border-color:var(--accent);background:var(--panel2)}
.card.active{border-color:var(--accent);box-shadow:0 0 0 1px var(--accent)}
.card .top{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.card .term{font-weight:700;font-size:15px}
.card .en{color:var(--muted);font-size:12px}
.card .d{color:var(--muted);font-size:13px;margin-top:4px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.imp{font-size:11px;font-weight:700;border-radius:6px;padding:1px 7px;color:#0b1220}
.imp.a{background:var(--imp-a)}.imp.b{background:var(--imp-b)}.imp.c{background:var(--imp-c)}
.cat{font-size:11px;color:var(--accent);background:rgba(56,189,248,.12);padding:1px 8px;border-radius:6px}
mark{background:#fde047;color:#1f2937;padding:0 1px;border-radius:2px}
.detail h2{margin:0 0 4px;font-size:20px;display:flex;align-items:center;gap:8px}
.detail .den{color:var(--muted);font-size:13px;margin-bottom:14px}
.detail dt{color:var(--accent);font-size:12px;margin-top:14px;font-weight:700}
.detail dd{margin:4px 0 0;font-size:14px}
.detail .empty{color:var(--muted);text-align:center;margin-top:40px}
.rel{display:flex;flex-wrap:wrap;gap:6px;margin-top:6px}
.rel button{background:var(--panel2);border:1px solid var(--line);color:var(--text);font-size:12px;padding:4px 9px;border-radius:7px;cursor:pointer}
.rel button:hover{border-color:var(--accent);color:var(--accent)}
.noresult{text-align:center;color:var(--muted);margin-top:50px}
@media(max-width:760px){.layout{flex-direction:column;height:auto}
  .detail{flex:none;border-left:none;border-top:1px solid var(--line)}}
</style>
</head>
<body>
<header>
  <h1><a href="index.html">G検定 用語集 検索</a> <span class="badge">__N__語</span></h1>
  <div class="controls">
    <input id="q" type="search" placeholder="用語・英語・説明で検索（例: 勾配、CNN、過学習／カンマ区切りで複数検索可）" autofocus>
    <select id="cat"><option value="">全カテゴリ</option></select>
    <div class="imps">
      <label><input type="checkbox" class="impf" value="◎" checked>◎</label>
      <label><input type="checkbox" class="impf" value="○" checked>○</label>
      <label><input type="checkbox" class="impf" value="△" checked>△</label>
    </div>
  </div>
</header>
<div class="count" id="count"></div>
<div class="layout">
  <div class="list" id="list"></div>
  <div class="detail" id="detail"><div class="empty">用語を選ぶと<br>ここに解説が出ます</div></div>
</div>
<script>
const DATA = __DATA__;
const CATS = __CATS__;
const impClass = i => i==="◎"?"a":i==="○"?"b":"c";
const $ = s => document.querySelector(s);
const q=$("#q"), catSel=$("#cat"), list=$("#list"), detail=$("#detail"), count=$("#count");
CATS.forEach(c=>{const o=document.createElement("option");o.value=c;o.textContent=c;catSel.appendChild(o);});
let activeId=null;
const esc=s=>(s||"").replace(/[&<>]/g,m=>({"&":"&amp;","<":"&lt;",">":"&gt;"}[m]));
function parseKeywords(raw){
  return [...new Set(raw.split(/[,、]/).map(s=>s.trim()).filter(Boolean))];
}
function hl(s,kws){s=esc(s);if(!kws||!kws.length)return s;
  const pat=kws.map(k=>k.replace(/[.*+?^${}()|[\]\\]/g,"\\$&")).join("|");
  try{return s.replace(new RegExp("("+pat+")","gi"),"<mark>$1</mark>");}catch(e){return s;}}
function activeImps(){return [...document.querySelectorAll(".impf:checked")].map(c=>c.value);}
function matchesAny(text,kws){const t=text.toLowerCase();return kws.some(k=>t.includes(k.toLowerCase()));}
function filtered(kws){
  const cat=catSel.value, imps=activeImps();
  return DATA.filter(d=>{
    if(cat&&d.cat!==cat)return false;
    if(!imps.includes(d.imp))return false;
    if(!kws.length)return true;
    return matchesAny(d.term+" "+d.en+" "+d.desc+" "+d.note, kws);
  });
}
function cardHtml(d,kws){return `
    <div class="card${d.id===activeId?' active':''}" data-id="${d.id}">
      <div class="top"><span class="imp ${impClass(d.imp)}">${d.imp}</span>
        <span class="term">${hl(d.term,kws)}</span>
        ${d.en?`<span class="en">${hl(d.en,kws)}</span>`:""}
        <span class="cat">${esc(d.cat)}</span></div>
      <div class="d">${hl(d.desc,kws)}</div></div>`;}
function render(){
  const raw=q.value.trim(), kws=parseKeywords(raw), rows=filtered(kws);
  const kwLabel = kws.length>1 ? kws.map(k=>`「${k}」`).join("・")+"のいずれか" : (kws[0]?`「${kws[0]}」`:"");
  count.textContent=rows.length+" 件"+(kwLabel?"（"+kwLabel+"で絞り込み）":"");
  let html = rows.map(d=>cardHtml(d,kws)).join("");
  list.innerHTML = html || '<div class="noresult">該当する用語がありません</div>';
}
function showDetail(id){
  activeId=id;
  const d=DATA.find(x=>x.id===id);if(!d)return;
  const rel=DATA.filter(x=>x.cat===d.cat&&x.id!==d.id).slice(0,8);
  detail.innerHTML=`
    <h2><span class="imp ${impClass(d.imp)}">${d.imp}</span> ${esc(d.term)}</h2>
    <div class="den">${esc(d.en)} ・ ${esc(d.cat)}</div>
    <dl><dt>意味・説明</dt><dd>${esc(d.desc)}</dd>
      ${d.note?`<dt>補足・関連</dt><dd>${esc(d.note)}</dd>`:""}
      <dt>同カテゴリの関連用語</dt>
      <dd><div class="rel">${rel.map(r=>`<button data-id="${r.id}">${esc(r.term)}</button>`).join("")||"—"}</div></dd>
    </dl>`;
  document.querySelectorAll(".card").forEach(c=>c.classList.toggle("active",+c.dataset.id===id));
  detail.querySelectorAll(".rel button").forEach(b=>b.onclick=()=>{showDetail(+b.dataset.id);detail.scrollTop=0;});
}
list.addEventListener("click",e=>{const c=e.target.closest(".card");if(c)showDetail(+c.dataset.id);});
q.addEventListener("input",render);
catSel.addEventListener("change",render);
document.querySelectorAll(".impf").forEach(c=>c.addEventListener("change",render));
document.addEventListener("keydown",e=>{
  if(e.key==="/"&&document.activeElement!==q){e.preventDefault();q.focus();}
  if(e.key==="Escape"){detail.innerHTML='<div class="empty">用語を選ぶと<br>ここに解説が出ます</div>';activeId=null;render();}
});
render();
</script>
</body>
</html>
"""

html = (HTML.replace("__DATA__", data_json)
            .replace("__CATS__", cats_json)
            .replace("__N__", str(len(records))))
with open(OUT, "w", encoding="utf-8") as f:
    f.write(html)
print("saved:", OUT, "/", len(records), "terms,", len(cats), "categories")
