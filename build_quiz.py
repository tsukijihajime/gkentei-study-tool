# -*- coding: utf-8 -*-
"""G検定 用語集 xlsx -> 4択クイズ 単一HTML（GitHub Pages公開版）を生成する。

出題形式：説明文 -> 該当する用語名を4択から選ぶ。
ダミー選択肢は同カテゴリ内の他の用語からランダムに選出する。
gkentei_glossary_public.xlsx（説明文を書き直した公開版データ）を読み込む。
"""
import json
import os
import openpyxl

XLSX = "gkentei_glossary_public.xlsx"
OUT = "quiz.html"

ws = openpyxl.load_workbook(XLSX).active
records = []
for r in range(2, ws.max_row + 1):
    no, cat, term, en, desc, note, imp, _ = [ws.cell(r, c).value for c in range(1, 9)]
    if not term:
        continue
    records.append({"id": no, "cat": cat or "", "term": term or "", "en": en or "",
                    "desc": desc or "", "note": note or "", "imp": imp or ""})

def catkey(x):
    head = x.split(".")[0]
    return float(head) if head.replace(".", "").isdigit() else 999.0

cats = sorted({r["cat"] for r in records}, key=catkey)
data_json = json.dumps(records, ensure_ascii=False)
cats_json = json.dumps(cats, ensure_ascii=False)

HTML = r"""<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>G検定 問題集</title>
<style>
:root{--bg:#0f172a;--panel:#1e293b;--panel2:#273449;--line:#334155;--text:#e2e8f0;
  --muted:#94a3b8;--accent:#38bdf8;--accent2:#0ea5e9;--imp-a:#f87171;--imp-b:#fbbf24;--imp-c:#34d399;
  --ok:#34d399;--ng:#f87171;}
*{box-sizing:border-box}
body{margin:0;font-family:"Meiryo UI","Segoe UI",system-ui,sans-serif;background:var(--bg);color:var(--text);line-height:1.6}
header{position:sticky;top:0;z-index:10;background:rgba(15,23,42,.96);backdrop-filter:blur(6px);
  border-bottom:1px solid var(--line);padding:14px 20px;display:flex;align-items:center;justify-content:space-between;gap:10px}
h1{margin:0;font-size:18px;display:flex;align-items:center;gap:10px}
h1 .badge{font-size:11px;background:var(--accent2);color:#02283a;padding:2px 8px;border-radius:10px;font-weight:700}
.timer{font-size:15px;font-weight:700;color:var(--accent);font-variant-numeric:tabular-nums;display:none}
main{max-width:720px;margin:0 auto;padding:24px 20px 60px}
.card{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:22px}
h2{margin:0 0 16px;font-size:16px;color:var(--accent)}
.grp{margin-bottom:20px}
.grp .label{font-size:13px;color:var(--muted);margin-bottom:8px;display:flex;justify-content:space-between;align-items:center}
.grp .label button{background:none;border:none;color:var(--accent);font-size:12px;cursor:pointer;padding:0}
.grp .label .labelbtns{display:flex;gap:10px}
.chips{display:flex;flex-wrap:wrap;gap:7px}
.chips label{display:flex;align-items:center;gap:5px;background:var(--panel2);border:1px solid var(--line);
  padding:6px 11px;border-radius:8px;font-size:13px;cursor:pointer;user-select:none}
.chips input{accent-color:var(--accent)}
select{padding:9px 12px;border-radius:9px;border:1px solid var(--line);background:var(--panel2);color:var(--text);font-size:14px}
.poolcount{color:var(--muted);font-size:13px;margin-top:4px}
.startbtn{margin-top:8px;width:100%;padding:13px;font-size:15px;font-weight:700;border-radius:10px;border:none;
  background:var(--accent2);color:#02283a;cursor:pointer}
.startbtn:disabled{background:var(--line);color:var(--muted);cursor:not-allowed}
.imp{font-size:11px;font-weight:700;border-radius:6px;padding:1px 7px;color:#0b1220}
.imp.a{background:var(--imp-a)}.imp.b{background:var(--imp-b)}.imp.c{background:var(--imp-c)}
.qhead{display:flex;justify-content:space-between;align-items:center;color:var(--muted);font-size:13px;margin-bottom:10px}
.qcat{font-size:11px;color:var(--accent);background:rgba(56,189,248,.12);padding:1px 8px;border-radius:6px}
.qdesc{font-size:16px;margin:10px 0 20px;padding:16px;background:var(--panel2);border-radius:10px;border:1px solid var(--line)}
.choices{display:flex;flex-direction:column;gap:9px}
.choice{text-align:left;padding:12px 15px;border-radius:9px;border:1px solid var(--line);background:var(--panel2);
  color:var(--text);font-size:15px;cursor:pointer;transition:.12s}
.choice:hover:not(:disabled){border-color:var(--accent)}
.choice:disabled{cursor:default}
.choice.correct{border-color:var(--ok);background:rgba(52,211,153,.15);color:#a7f3d0}
.choice.wrong{border-color:var(--ng);background:rgba(248,113,113,.15);color:#fecaca}
.result{margin-top:18px;padding:16px;border-radius:10px;border:1px dashed var(--line)}
.result .verdict{font-weight:700;font-size:15px;margin-bottom:8px}
.result .verdict.ok{color:var(--ok)}
.result .verdict.ng{color:var(--ng)}
.result dt{color:var(--accent);font-size:12px;margin-top:10px;font-weight:700}
.result dd{margin:4px 0 0;font-size:14px}
.nextbtn{margin-top:16px;width:100%;padding:12px;font-size:14px;font-weight:700;border-radius:9px;border:none;
  background:var(--accent2);color:#02283a;cursor:pointer}
.scorebig{font-size:44px;font-weight:700;text-align:center;margin:10px 0}
.scoresub{text-align:center;color:var(--muted);margin-bottom:8px}
.timeinfo{display:flex;justify-content:center;gap:22px;color:var(--muted);font-size:13px;margin-bottom:4px}
.timeinfo b{color:var(--text)}
.pace{text-align:center;font-size:13px;margin-bottom:22px}
.pace.ok{color:var(--ok)}
.pace.ng{color:var(--ng)}
.pace b{font-weight:700}
.resultbtns{display:flex;gap:10px}
.resultbtns button{flex:1;padding:12px;border-radius:9px;font-size:14px;font-weight:700;cursor:pointer;border:1px solid var(--line)}
.resultbtns .again{background:var(--accent2);color:#02283a;border:none}
.resultbtns .back{background:var(--panel2);color:var(--text)}
.progressbar{height:6px;border-radius:3px;background:var(--line);overflow:hidden;margin-bottom:18px}
.progressbar .fill{height:100%;background:var(--accent2);transition:width .2s}
</style>
</head>
<body>
<header><h1>G検定 問題集 <span class="badge">__N__語</span></h1><span class="timer" id="timer">00:00</span></header>
<main id="app"></main>
<script>
const DATA = __DATA__;
const CATS = __CATS__;
const app = document.getElementById("app");
const esc = s => (s||"").replace(/[&<>]/g, m => ({"&":"&amp;","<":"&lt;",">":"&gt;"}[m]));
const impClass = i => i==="◎"?"a":i==="○"?"b":"c";
function shuffle(arr){
  const a = arr.slice();
  for(let i=a.length-1;i>0;i--){const j=Math.floor(Math.random()*(i+1));[a[i],a[j]]=[a[j],a[i]];}
  return a;
}
let state = null;
let timerInterval = null;
const timerEl = document.getElementById("timer");
function fmtTime(sec){
  const m = Math.floor(sec/60), s = sec%60;
  return String(m).padStart(2,"0") + ":" + String(s).padStart(2,"0");
}
function tickTimer(){
  timerEl.textContent = fmtTime(Math.floor((Date.now()-state.quizStart)/1000));
}
function startTimer(){
  timerEl.style.display = "inline";
  tickTimer();
  timerInterval = setInterval(tickTimer, 500);
}
function stopTimer(){
  if(timerInterval){ clearInterval(timerInterval); timerInterval = null; }
}

function renderSetup(){
  stopTimer();
  timerEl.style.display = "none";
  const catChips = CATS.map(c=>`<label><input type="checkbox" class="catf" value="${esc(c)}" checked>${esc(c)}</label>`).join("");
  app.innerHTML = `
    <div class="card">
      <h2>出題範囲を選んで開始</h2>
      <div class="grp">
        <div class="label"><span>カテゴリ</span><span class="labelbtns"><button id="catAll">全選択</button><button id="catNone">全解除</button></span></div>
        <div class="chips">${catChips}</div>
      </div>
      <div class="grp">
        <div class="label"><span>重要度</span></div>
        <div class="chips">
          <label><input type="checkbox" class="impf" value="◎" checked>◎</label>
          <label><input type="checkbox" class="impf" value="○" checked>○</label>
          <label><input type="checkbox" class="impf" value="△" checked>△</label>
        </div>
      </div>
      <div class="grp">
        <div class="label"><span>問題数</span></div>
        <select id="qcount">
          <option value="10">10問</option>
          <option value="20">20問</option>
          <option value="30">30問</option>
          <option value="all">全問</option>
        </select>
      </div>
      <div class="poolcount" id="poolcount"></div>
      <button class="startbtn" id="startBtn">開始</button>
    </div>`;
  const catAll = document.getElementById("catAll");
  const catNone = document.getElementById("catNone");
  const catBoxes = () => [...document.querySelectorAll(".catf")];
  const impBoxes = () => [...document.querySelectorAll(".impf")];
  function pool(){
    const cats = catBoxes().filter(c=>c.checked).map(c=>c.value);
    const imps = impBoxes().filter(c=>c.checked).map(c=>c.value);
    return DATA.filter(d=>cats.includes(d.cat) && imps.includes(d.imp));
  }
  function updateCount(){
    const p = pool();
    document.getElementById("poolcount").textContent = p.length + " 語が対象";
    document.getElementById("startBtn").disabled = p.length < 4;
  }
  catAll.onclick = () => { catBoxes().forEach(c=>c.checked=true); updateCount(); };
  catNone.onclick = () => { catBoxes().forEach(c=>c.checked=false); updateCount(); };
  catBoxes().forEach(c=>c.addEventListener("change", updateCount));
  impBoxes().forEach(c=>c.addEventListener("change", updateCount));
  updateCount();
  document.getElementById("startBtn").onclick = () => {
    const p = pool();
    const countSel = document.getElementById("qcount").value;
    const n = countSel==="all" ? p.length : Math.min(parseInt(countSel,10), p.length);
    const questions = shuffle(p).slice(0, n);
    state = { questions, idx: 0, correct: 0, answered: false, quizStart: Date.now() };
    startTimer();
    renderQuestion();
  };
}

function makeChoices(correct){
  const sameCat = DATA.filter(d=>d.cat===correct.cat && d.id!==correct.id);
  const others = DATA.filter(d=>d.id!==correct.id && d.cat!==correct.cat);
  const dummies = shuffle(sameCat).slice(0,3);
  if(dummies.length<3) dummies.push(...shuffle(others).slice(0, 3-dummies.length));
  return shuffle([correct, ...dummies]);
}

function renderQuestion(){
  const { questions, idx } = state;
  const q = questions[idx];
  const choices = makeChoices(q);
  state.answered = false;
  const pct = Math.round(idx/questions.length*100);
  app.innerHTML = `
    <div class="card">
      <div class="progressbar"><div class="fill" style="width:${pct}%"></div></div>
      <div class="qhead"><span>第 ${idx+1} 問 / ${questions.length} 問</span><span class="qcat">${esc(q.cat)}</span></div>
      <div class="qdesc">${esc(q.desc)}</div>
      <div class="choices">${choices.map((c,i)=>`<button class="choice" data-id="${c.id}">${esc(c.term)}${c.en?` <span style="color:var(--muted);font-size:12px">(${esc(c.en)})</span>`:""}</button>`).join("")}</div>
      <div id="resultArea"></div>
    </div>`;
  document.querySelectorAll(".choice").forEach(btn=>{
    btn.onclick = () => onAnswer(+btn.dataset.id, q);
  });
}

function onAnswer(pickedId, q){
  if(state.answered) return;
  state.answered = true;
  const ok = pickedId === q.id;
  if(ok) state.correct++;
  document.querySelectorAll(".choice").forEach(btn=>{
    btn.disabled = true;
    const id = +btn.dataset.id;
    if(id===q.id) btn.classList.add("correct");
    else if(id===pickedId) btn.classList.add("wrong");
  });
  const resultArea = document.getElementById("resultArea");
  const isLast = state.idx === state.questions.length-1;
  resultArea.innerHTML = `
    <div class="result">
      <div class="verdict ${ok?'ok':'ng'}">${ok?'○ 正解':'✕ 不正解'}</div>
      <dt>正解：<span class="imp ${impClass(q.imp)}">${esc(q.imp)}</span> ${esc(q.term)} ${q.en?`(${esc(q.en)})`:""}</dt>
      <dd>${esc(q.desc)}</dd>
      ${q.note?`<dt>補足・関連</dt><dd>${esc(q.note)}</dd>`:""}
    </div>
    <button class="nextbtn" id="nextBtn">${isLast?'結果を見る':'次の問題へ'}</button>`;
  document.getElementById("nextBtn").onclick = () => {
    if(isLast){ renderResult(); } else { state.idx++; renderQuestion(); }
  };
}

const EXAM_PACE_SEC = 40;
function renderResult(){
  stopTimer();
  const { questions, correct } = state;
  const pct = Math.round(correct/questions.length*100);
  const totalSec = Math.floor((Date.now()-state.quizStart)/1000);
  const avgSec = totalSec / questions.length;
  const diff = EXAM_PACE_SEC - avgSec;
  const paceOk = diff >= 0;
  const paceMsg = paceOk
    ? `本番想定ペース（${EXAM_PACE_SEC}秒/問）より <b>${diff.toFixed(1)}秒</b> 速いペース`
    : `本番想定ペース（${EXAM_PACE_SEC}秒/問）より <b>${Math.abs(diff).toFixed(1)}秒</b> 遅いペース`;
  app.innerHTML = `
    <div class="card">
      <h2>結果</h2>
      <div class="scorebig">${correct} / ${questions.length}</div>
      <div class="scoresub">正答率 ${pct}%</div>
      <div class="timeinfo"><span>合計時間 <b>${fmtTime(totalSec)}</b></span><span>平均解答時間 <b>${avgSec.toFixed(1)}秒/問</b></span></div>
      <div class="pace ${paceOk?'ok':'ng'}">${paceMsg}</div>
      <div class="resultbtns">
        <button class="again" id="againBtn">もう一度</button>
        <button class="back" id="backBtn">設定に戻る</button>
      </div>
    </div>`;
  document.getElementById("backBtn").onclick = renderSetup;
  document.getElementById("againBtn").onclick = () => {
    state = { questions: shuffle(questions), idx: 0, correct: 0, answered: false, quizStart: Date.now() };
    startTimer();
    renderQuestion();
  };
}

renderSetup();
</script>
</body>
</html>
"""

html = (HTML.replace("__DATA__", data_json)
            .replace("__CATS__", cats_json)
            .replace("__N__", str(len(records))))
with open(OUT, "w", encoding="utf-8") as f:
    f.write(html)
print("saved:", OUT, "/", len(records), "terms,", len(cats), "categories")
